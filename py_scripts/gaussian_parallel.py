
#%%
import matplotlib.pyplot as plt
from skimage import io, color
from skimage.feature import blob_log
import os
import pandas as pd
import numpy as np
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font
from joblib import Parallel, delayed
import random
import gc

# ----------------------------
# Parameters
# ----------------------------
IMAGE_DIR = "../images/CCSS/"
OUTPUT_CSV = "./image_segmentation_grey.csv"
OUTPUT_XLSX = "./boat_detections_parallel.xlsx"
OUTPUT_IMG_DIR = "./output_parallel/"
MIN_SIGMA = 2
MAX_SIGMA = 10
NUM_SIGMA = 5
THRESHOLD = 0.3
TOP_PERCENT = 0.20
BOTTOM_PERCENT = 0.20
BATCH_SIZE = 3000
N_JOBS = 4  # number of parallel processes

os.makedirs(OUTPUT_IMG_DIR, exist_ok=True)

# ----------------------------
# Load and sample image list
# ----------------------------
df = pd.read_csv(OUTPUT_CSV)
grey_images = df[df["grey"]]["image"].tolist()

random.seed(42)
#grey_images = random.sample(grey_images_full, min(5000, len(grey_images_full)))

# ----------------------------
# Image processing function
# ----------------------------
def process_single_image(img_name):
    try:
        img_path = os.path.join(IMAGE_DIR, img_name)
        img = io.imread(img_path)
        gray_img = color.rgb2gray(img)
        height, width = gray_img.shape

        # Crop region of interest
        top_cut = int(height * TOP_PERCENT)
        bottom_cut = int(height * (1 - BOTTOM_PERCENT))
        gray = gray_img[top_cut:bottom_cut, :]

        # Detect blobs (boats)
        blobs = blob_log(gray, min_sigma=MIN_SIGMA, max_sigma=MAX_SIGMA,
                         num_sigma=NUM_SIGMA, threshold=THRESHOLD)
        num_boats = len(blobs)

        # Save visualization
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.imshow(gray, cmap='gray')
        for y, x, sigma in blobs:
            c = plt.Circle((x, y), sigma * 1.5, color='red', linewidth=1.5, fill=False)
            ax.add_patch(c)
        ax.set_title(f"{img_name} — Boats: {num_boats}")
        ax.axis('off')

        output_img_path = os.path.join(OUTPUT_IMG_DIR, f"{os.path.splitext(img_name)[0]}_detected.jpg")
        plt.savefig(output_img_path, bbox_inches='tight', pad_inches=0)
        plt.close(fig)

        blob_coords = "; ".join([f"({x:.1f}, {y:.1f})" for y, x, sigma in blobs])

        return {
            "image": img_name,
            "boat_count": num_boats,
            "detected_image": os.path.abspath(output_img_path),
            "boat_coordinates": blob_coords
        }

    except Exception as e:
        return {
            "image": img_name,
            "boat_count": 0,
            "detected_image": None,
            "boat_coordinates": f"Error: {e}"
        }

# ----------------------------
# Function to append batch to Excel
# ----------------------------
def append_to_excel(df_batch, filename):
    if not os.path.exists(filename):
        df_batch.to_excel(filename, index=False)
    else:
        with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            start_row = writer.sheets["Sheet1"].max_row
            df_batch.to_excel(writer, index=False, header=False, startrow=start_row)

# ----------------------------
# Main processing loop
# ----------------------------
pbar = tqdm(total=len(grey_images), desc="Processing images", ncols=80)

for i in range(0, len(grey_images), BATCH_SIZE):
    batch = grey_images[i:i + BATCH_SIZE]
    print(f"\n🧩 Processing batch {i // BATCH_SIZE + 1} ({len(batch)} images)")

    # Parallel image processing (separate processes = better memory cleanup)
    batch_results = Parallel(n_jobs=N_JOBS, prefer="processes")(
        delayed(process_single_image)(img_name) for img_name in batch
    )

    # Create DataFrame and append to Excel
    batch_df = pd.DataFrame(batch_results)
    append_to_excel(batch_df, OUTPUT_XLSX)

    # Update progress bar
    pbar.update(len(batch))

    # Release memory between batches
    del batch_results, batch_df
    gc.collect()
    plt.close('all')

pbar.close()

# ----------------------------
# Add clickable hyperlinks
# ----------------------------
wb = load_workbook(OUTPUT_XLSX)
ws = wb.active

col_idx = None
for i, cell in enumerate(ws[1], start=1):
    if cell.value == "detected_image":
        col_idx = i
        break

if col_idx is not None:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        img_path = cell.value
        if img_path and os.path.exists(img_path):
            cell.value = "View Image"
            cell.hyperlink = img_path
            cell.font = Font(color="0000FF", underline="single")

wb.save(OUTPUT_XLSX)

print(f"\n✅ Saved all boat detections to {OUTPUT_XLSX}")
print(f"✅ Output images in: {OUTPUT_IMG_DIR}")
#%%
