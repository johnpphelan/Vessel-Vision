#%%
# preserve_hyperlinks_sample.py
import random
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# ----------------------------
# Parameters
# ----------------------------
INPUT_XLSX = "./boat_detections_parallel.xlsx"
OUTPUT_SAMPLE_XLSX = "./boat_detections_sample_preserved.xlsx"
SAMPLE_SIZE = 100
RANDOM_SEED = 42

# ----------------------------
# Load original workbook
# ----------------------------
wb = load_workbook(INPUT_XLSX, read_only=False, data_only=False)
ws = wb.active  # assume data is on the first sheet

max_row = ws.max_row
max_col = ws.max_column

# Determine which rows are available to sample (exclude header row 1)
available_rows = list(range(2, max_row + 1))
sample_size = min(SAMPLE_SIZE, len(available_rows))

random.seed(RANDOM_SEED)
sample_rows = sorted(random.sample(available_rows, sample_size))

# ----------------------------
# Create new workbook and copy header
# ----------------------------
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = ws.title + "_sample"

# Copy header row (row 1)
for col in range(1, max_col + 1):
    src_cell = ws.cell(row=1, column=col)
    dst_cell = ws_out.cell(row=1, column=col, value=src_cell.value)
    # copy basic font if present
    if src_cell.font:
        try:
            dst_cell.font = src_cell.font
        except Exception:
            pass

# ----------------------------
# Copy sampled rows (preserving hyperlinks)
# ----------------------------
out_row_idx = 2
for src_row in sample_rows:
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws_out.cell(row=out_row_idx, column=col, value=src_cell.value)

        # If the source cell has a hyperlink object, preserve it
        if src_cell.hyperlink is not None and src_cell.hyperlink.target:
            # set hyperlink target and display text
            # keep the cell.value if it's user-friendly (e.g., "View Image"), else set a generic label
            display_text = src_cell.value if src_cell.value else "View Image"
            dst_cell.value = display_text
            dst_cell.hyperlink = src_cell.hyperlink.target
            # style to look like a link
            dst_cell.font = Font(color="0000FF", underline="single")
        else:
            # If original cell had a font, try to copy (best-effort)
            if src_cell.font:
                try:
                    dst_cell.font = src_cell.font
                except Exception:
                    pass
    out_row_idx += 1

# ----------------------------
# Save sampled workbook
# ----------------------------
wb_out.save(OUTPUT_SAMPLE_XLSX)
print(f"Saved sample of {sample_size} rows to: {OUTPUT_SAMPLE_XLSX}")

#%%