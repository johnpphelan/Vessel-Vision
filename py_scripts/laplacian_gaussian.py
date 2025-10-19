#%%
#%%
import matplotlib.pyplot as plt
from skimage import io, color
from skimage.feature import blob_log
import os
import pandas as pd
import numpy as np
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font

# ----------------------------
# Parameters
# ----------------------------
IMAGE_DIR = "../images/CCSS/"
OUTPUT_CSV = "./image_segmentation_grey.csv"
OUTPUT_XLSX = "./boat_detections.xlsx"
OUTPUT_IMG_DIR = "./output/"
MIN_SIGMA = 2
MAX_SIGMA = 10
NUM_SIGMA = 5
THRESHOLD = 0.3
TOP_PERCENT = 0.25
BOTTOM_PERCENT = 0.20
BATCH_SIZE = 1000

# ----------------------------
# Setup
# ----------------------------
os.makedirs(OUTPUT_IMG_DIR, exist_ok=True)

# Load image list
df = pd.read_csv(OUTPUT_CSV)
# get the first 5000 images marked as grey
grey_images = df[df["grey"]]["image"].tolist()[:5000]

#%%
all_results = []

# ----------------------------
# Process images in batches
# ----------------------------
num_batches = (len(grey_images) + BATCH_SIZE - 1) // BATCH_SIZE

for batch_idx in range(num_batches):
    batch_images = grey_images[batch_idx * BATCH_SIZE : (batch_idx + 1) * BATCH_SIZE]
    
    print(f"\nProcessing batch {batch_idx + 1}/{num_batches} ({len(batch_images)} images)")
    
    batch_results = []
    
    for i, img_name in enumerate(tqdm(batch_images, desc="Processing images", ncols=80)):
        img_path = os.path.join(IMAGE_DIR, img_name)
        
        # Load and convert
        img = io.imread(img_path)
        gray_img = color.rgb2gray(img)
        height, width = gray_img.shape

        # Crop ROI
        top_cut = int(height * TOP_PERCENT)
        bottom_cut = int(height * (1 - BOTTOM_PERCENT))
        gray = gray_img[top_cut:bottom_cut, :]

        # Detect blobs (boats)
        blobs = blob_log(gray, min_sigma=MIN_SIGMA, max_sigma=MAX_SIGMA,
                         num_sigma=NUM_SIGMA, threshold=THRESHOLD)
        num_boats = len(blobs)

        # Create and save visualization
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.imshow(gray, cmap='gray')
        for y, x, sigma in blobs:
            c = plt.Circle((x, y), sigma * 1.5, color='red', linewidth=1.5, fill=False)
            ax.add_patch(c)
        ax.set_title(f"{img_name} — Boats: {num_boats}")
        ax.axis('off')

        output_img_path = os.path.join(OUTPUT_IMG_DIR, f"{os.path.splitext(img_name)[0]}_detected.jpg")
        plt.savefig(output_img_path, bbox_inches='tight', pad_inches=0)
        plt.close(fig)

        # Store data
        blob_coords = "; ".join([f"({x:.1f}, {y:.1f})" for y, x, sigma in blobs])
        batch_results.append({
            "image": img_name,
            "boat_count": num_boats,
            "detected_image": os.path.abspath(output_img_path),  # full path for hyperlink
            "boat_coordinates": blob_coords
        })
    
    # Append batch results
    all_results.extend(batch_results)

# ----------------------------
# Save to Excel
# ----------------------------
boat_counts_df = pd.DataFrame(all_results)
boat_counts_df.to_excel(OUTPUT_XLSX, index=False)

# ----------------------------
# Add clickable hyperlinks
# ----------------------------
wb = load_workbook(OUTPUT_XLSX)
ws = wb.active

# Find the column index for "detected_image"
col_idx = None
for i, cell in enumerate(ws[1], start=1):
    if cell.value == "detected_image":
        col_idx = i
        break

if col_idx is not None:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        img_path = cell.value
        cell.value = "View Image"
        cell.hyperlink = img_path
        cell.font = Font(color="0000FF", underline="single")

wb.save(OUTPUT_XLSX)

print(f"\n✅ Saved boat detections to {OUTPUT_XLSX}")
print(f"✅ Saved output images to {OUTPUT_IMG_DIR}")




#%%